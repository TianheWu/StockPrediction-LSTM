import torch
import torch.utils.data
import numpy as np

from openpyxl import load_workbook
from utils.data_process.process_data import get_sheet_col, normalize


class StockDataset(torch.utils.data.Dataset):
    """Dataset of stock"""
    def __init__(self, path_data, start_idx=0.0, end_idx=1.0, pre=4):
        wb = load_workbook(path_data)
        sheet = wb.active
        col_num = sheet.max_column
        dataset = []
        for i in range(1, col_num):
            if i == 7:
                continue
            col = get_sheet_col(sheet, i)[1:-2]
            dataset.append(normalize(col))
            if i == 1:
                start = np.array(normalize(col)[1:])
            if i == 4:
                ret_label = np.array(normalize(col)[1:])

        dataset = np.array(dataset)
        v1 = dataset[0][:, np.newaxis]
        v2 = dataset[1][:, np.newaxis]
        ret_x = np.concatenate((v1, v2), axis=1)
        for i in range(2, col_num - 2):
            v = dataset[i][:, np.newaxis]
            ret_x = np.concatenate((ret_x, v), axis=1)
        
        ret_x = ret_x[:-1]
        ret_x = np.concatenate((ret_x, start[:, np.newaxis]), axis=1)

        len_data = ret_x.shape[0]
        ret_x = ret_x[int(start_idx * len_data):int(end_idx * len_data)]
        ret_label = ret_label[int(start_idx * len_data):int(end_idx * len_data)]
        len_data = int(end_idx * len_data) - int(start_idx * len_data)

        ret_data = []
        label = []
        idx = 0
        while idx < len_data:
            if idx + pre >= len_data:
                break
            tmp_data = []
            for i in range(pre):
                tmp_data.append(ret_x[i + idx])
            idx += pre
            label.append(ret_label[idx])
            ret_data.append(tmp_data)
        
        ret_data = np.array(ret_data)
        label = np.array(label)

        self.data = torch.from_numpy(ret_data).type(torch.double)
        self.label = torch.from_numpy(label).type(torch.double)

    def __getitem__(self, index):
        return self.data[index], self.label[index]

    def __len__(self):
        return len(self.data)
    